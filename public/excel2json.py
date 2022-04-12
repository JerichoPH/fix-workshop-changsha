import json
import os

from openpyxl import load_workbook

from ComplexEncoder import ComplexEncoder


class FileNotFoundException(Exception):
    pass


def execl2json(file_path: str):
    """
    Excel转JSON格式
    :param file_path:
    :return:
    """
    if os.path.exists('%s.xlsx' % (file_path,)) is False:
        raise FileNotFoundException('文件不存在：%s' % (file_path,))

    excel_reader = load_workbook(filename=file_path + ".xlsx")
    sheet_names = excel_reader.sheetnames
    data = {}
    for sheet_name in sheet_names:
        sheet = excel_reader[sheet_name]
        sheet_data = []
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row in range(4, max_row + 1):
            row_data = []
            for column in range(1, max_column + 1):
                cell = sheet.cell(row=row, column=column).value
                row_data.append(cell)
            sheet_data.append(row_data)
        data[sheet_name] = sheet_data

    json.dump(data, open(file_path + ".json", "w", encoding="utf-8"), ensure_ascii=False, cls=ComplexEncoder)

    print('执行完成：%s' % (file_path,))


file_paths = [
    "C:\\xampp\\htdocs\\maintain\\public\\转辙机"
]

try:
    for file_path in file_paths:
        execl2json(file_path=file_path)
except FileNotFoundException as e:
    print(e)
