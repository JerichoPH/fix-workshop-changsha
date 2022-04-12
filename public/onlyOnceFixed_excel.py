import os
import sys
from os import system

from openpyxl import Workbook, load_workbook


def save_excel(filename: str, excel_path: str, new_excel: Workbook):
    """
    读取旧Excel组合新Excel
    :param filename:
    :param excel_path:
    :param year:
    :param month:
    :param new_excel:
    :return:
    """
    # 种类统计
    new_sheet = new_excel.create_sheet('%s统计' % (filename,))
    if os.path.exists("%s/%s" % (excel_path, '%s.xlsx' % (filename,),)) is False:
        print('文件不存在')
        exit(0)
    else:
        excel_reader = load_workbook(filename="%s/%s" % (excel_path, '%s.xlsx' % (filename,),))
        sheet_names = excel_reader.sheetnames
        for sheet_name in sheet_names:
            for row in excel_reader[sheet_name].rows:
                new_sheet.append((row[0].value, row[1].value, row[2].value, row[3].value,))


excel_path = sys.argv[1]

new_excel = Workbook()
new_excel.remove(new_excel['Sheet'])
save_excel(filename='种类', excel_path=excel_path, new_excel=new_excel)
save_excel(filename='类型', excel_path=excel_path, new_excel=new_excel)
save_excel(filename='型号和子类', excel_path=excel_path, new_excel=new_excel)
ret = new_excel.save('%s/%s' % (excel_path, '统计.xlsx'))
print(ret)

exit(0)
